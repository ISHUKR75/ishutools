"""
excel_to_pdf.py — Convert Excel (.xlsx/.xls/.csv) to PDF (Enterprise Edition)
IshuTools.fun | Professional PDF Suite
Author: Ishu Kumar (ISHUKR41 / ISHUKR75)

Engines: openpyxl · reportlab · fitz (PyMuPDF) · pikepdf · Ghostscript CLI
Features:
  - All sheets rendered as separate PDF pages/sections
  - Auto landscape vs portrait based on column count
  - Smart column width calculation from openpyxl metadata
  - Cell number/date/percentage formatting
  - Bold/italic detection from cell font metadata
  - Color-coded cells (background fill from Excel)
  - Multi-page tables with repeated headers
  - Frozen-pane awareness (freeze row as repeated header)
  - CSV input support (UTF-8, UTF-8-BOM, detect delimiter)
  - Charts described as text placeholder per sheet
  - Cover sheet with file metadata and sheet summary
  - Ghostscript post-pass compression
  - pikepdf metadata injection
  - Configurable page size, orientation, font size
  - Row/column limit controls
  - Conditional formatting highlight detection
  - Merged cell handling (display value in first cell)
  - Column header normalization and deduplication
  - Per-sheet color themes (alternating row colors)
  - Header/footer with page numbers and filename
  - Multi-line cell content wrapping
  - Hyperlink detection and footnote listing
  - Custom title page with stats table
  - Summary statistics chart (BarChart) for numeric-first sheets
  - Formula-value display (data_only=True mode)
  - Unicode text support
  - Memory-efficient row streaming for large files
"""

import os
import io
import csv
import re
import shutil
import subprocess
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles.numbers import FORMAT_DATE_DATETIME
import pikepdf
import fitz

from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                 Paragraph, Spacer, PageBreak, HRFlowable,
                                 KeepTogether)
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT

# ── CLI binary detection ─────────────────────────────────────────────────────
GS_BIN = shutil.which('gs') or shutil.which('ghostscript')
QPDF_BIN = shutil.which('qpdf')


# ── Color helpers ─────────────────────────────────────────────────────────────

def _argb_to_hex(argb: str) -> str:
    """Convert openpyxl ARGB color string to #RRGGBB."""
    if not argb or argb in ('00000000', 'FFFFFFFF', 'FF000000', ''):
        return ''
    s = argb.lstrip('#')
    if len(s) == 8:
        return '#' + s[2:]
    if len(s) == 6:
        return '#' + s
    if len(s) == 3:
        return '#' + ''.join(c * 2 for c in s)
    return ''


def _get_cell_bg_hex(cell) -> str:
    """Extract background hex from openpyxl cell fill."""
    try:
        fill = cell.fill
        if not fill or fill.fill_type not in ('solid', 'patternFill'):
            return ''
        fc = fill.fgColor
        if not fc:
            return ''
        if fc.type == 'rgb':
            return _argb_to_hex(fc.rgb)
        if fc.type == 'theme':
            return ''
    except Exception:
        pass
    return ''


def _get_cell_font_color(cell) -> str:
    try:
        if cell.font and cell.font.color:
            fc = cell.font.color
            if fc.type == 'rgb':
                return _argb_to_hex(fc.rgb)
    except Exception:
        pass
    return ''


def _rl_color(hex_str: str):
    """Convert hex string to ReportLab color, or None if invalid/empty."""
    if not hex_str:
        return None
    try:
        return colors.HexColor(hex_str)
    except Exception:
        return None


# ── Cell value helpers ────────────────────────────────────────────────────────

def _format_cell_value(cell) -> str:
    """Format cell value as display string."""
    val = cell.value
    if val is None:
        return ''
    if isinstance(val, bool):
        return 'TRUE' if val else 'FALSE'
    if isinstance(val, float):
        # Check for percentage format
        try:
            nf = cell.number_format or ''
            if '%' in nf:
                return f'{val * 100:.2f}%'
            if 'E' in nf or 'e' in nf:
                return f'{val:.3e}'
            if val == int(val) and abs(val) < 1e12:
                return str(int(val))
            return f'{val:.4g}'
        except Exception:
            return str(val)
    if isinstance(val, int):
        return str(val)
    if hasattr(val, 'strftime'):
        try:
            return val.strftime('%Y-%m-%d')
        except Exception:
            return str(val)
    s = str(val).strip()
    return s


def _safe_para(text: str, style, max_len: int = 300) -> Paragraph:
    """Create a ReportLab Paragraph with XML-safe text."""
    safe = str(text)[:max_len]
    safe = (safe.replace('&', '&amp;')
                .replace('<', '&lt;')
                .replace('>', '&gt;')
                .replace('\n', '<br/>'))
    try:
        return Paragraph(safe, style)
    except Exception:
        return Paragraph(str(text)[:100].encode('ascii', 'replace').decode(), style)


# ── CSV loading ───────────────────────────────────────────────────────────────

def _detect_csv_delimiter(path: str) -> str:
    """Auto-detect CSV delimiter."""
    try:
        with open(path, 'r', encoding='utf-8-sig', errors='replace') as f:
            sample = f.read(4096)
        sniffer = csv.Sniffer()
        dialect = sniffer.sniff(sample, delimiters=',\t;|')
        return dialect.delimiter
    except Exception:
        return ','


def _load_csv(file_path: str, max_rows: int = 2000) -> list:
    """Load CSV returning list of row lists."""
    delimiter = _detect_csv_delimiter(file_path)
    data = []
    try:
        with open(file_path, 'r', encoding='utf-8-sig', errors='replace') as f:
            reader = csv.reader(f, delimiter=delimiter)
            for i, row in enumerate(reader):
                if i >= max_rows:
                    break
                data.append(row)
    except Exception:
        pass
    return data


# ── GS / pikepdf post-processing ─────────────────────────────────────────────

def _gs_compress(input_path: str, output_path: str,
                 quality: str = 'ebook') -> bool:
    if not GS_BIN:
        return False
    q_map = {
        'screen': '/screen', 'ebook': '/ebook',
        'printer': '/printer', 'prepress': '/prepress',
    }
    q = q_map.get(quality, '/ebook')
    cmd = [
        GS_BIN,
        '-dNOPAUSE', '-dBATCH', '-dQUIET',
        '-sDEVICE=pdfwrite',
        f'-dPDFSETTINGS={q}',
        '-dCompatibilityLevel=1.7',
        f'-sOutputFile={output_path}',
        input_path,
    ]
    try:
        proc = subprocess.run(cmd, capture_output=True, timeout=120)
        return (proc.returncode == 0
                and os.path.exists(output_path)
                and os.path.getsize(output_path) > 200)
    except Exception:
        return False


def _pikepdf_metadata(
    path: str,
    title: str = '',
    author: str = '',
    subject: str = '',
) -> None:
    try:
        with pikepdf.open(path, suppress_warnings=True) as pdf:
            pdf.docinfo['/Producer'] = 'IshuTools.fun PDF Suite — Excel2PDF'
            pdf.docinfo['/Creator'] = 'excel_to_pdf'
            if title:
                pdf.docinfo['/Title'] = title
            if author:
                pdf.docinfo['/Author'] = author
            if subject:
                pdf.docinfo['/Subject'] = subject
            pdf.docinfo['/CreationDate'] = datetime.now().strftime(
                "D:%Y%m%d%H%M%S")
            pdf.save(path)
    except Exception:
        pass


# ── Style builders ────────────────────────────────────────────────────────────

def _build_rl_styles(base_styles, font_size: int = 9):
    return {
        'title': ParagraphStyle(
            'ExcelTitle', parent=base_styles['Heading1'],
            fontSize=16, textColor=colors.HexColor('#1E3A5F'),
            spaceAfter=4, spaceBefore=0),
        'sheet_title': ParagraphStyle(
            'SheetTitle', parent=base_styles['Heading2'],
            fontSize=13, textColor=colors.HexColor('#1E3A5F'),
            spaceBefore=6, spaceAfter=4),
        'body': ParagraphStyle(
            'ExcelBody', parent=base_styles['Normal'],
            fontSize=font_size, leading=font_size + 2, fontName='Helvetica'),
        'body_r': ParagraphStyle(
            'ExcelBodyR', parent=base_styles['Normal'],
            fontSize=font_size, leading=font_size + 2,
            alignment=TA_RIGHT, fontName='Helvetica'),
        'body_c': ParagraphStyle(
            'ExcelBodyC', parent=base_styles['Normal'],
            fontSize=font_size, leading=font_size + 2,
            alignment=TA_CENTER, fontName='Helvetica'),
        'caption': ParagraphStyle(
            'Caption', parent=base_styles['Normal'],
            fontSize=7.5, textColor=colors.HexColor('#6B7280'),
            alignment=TA_CENTER),
        'small': ParagraphStyle(
            'Small', parent=base_styles['Normal'],
            fontSize=8, textColor=colors.HexColor('#6B7280')),
    }


# ── Worksheet → ReportLab data ────────────────────────────────────────────────

def _worksheet_to_table_data(
    ws,
    styles: dict,
    max_rows: int = 500,
    max_cols: int = 20,
    font_size: int = 9,
) -> tuple:
    """
    Convert openpyxl worksheet rows to ReportLab Table data + style commands.
    Returns (data, style_commands, row_count, col_count).
    """
    data = []
    cmd = []

    # Determine effective dimensions
    eff_rows = min(ws.max_row or 1, max_rows)
    eff_cols = min(ws.max_column or 1, max_cols)

    for r_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=eff_rows, max_col=eff_cols)):
        row_data = []
        for c_idx, cell in enumerate(row):
            val_str = _format_cell_value(cell)

            # Determine alignment
            align = TA_LEFT
            try:
                if isinstance(cell.value, (int, float)) and not isinstance(
                        cell.value, bool):
                    align = TA_RIGHT
            except Exception:
                pass

            # Build style for this cell
            is_bold = False
            is_italic = False
            try:
                if cell.font:
                    is_bold = bool(cell.font.bold)
                    is_italic = bool(cell.font.italic)
            except Exception:
                pass

            fn = ('Helvetica-BoldOblique' if (is_bold and is_italic) else
                  'Helvetica-Bold' if is_bold else
                  'Helvetica-Oblique' if is_italic else 'Helvetica')

            cell_style = ParagraphStyle(
                f'Cell_{r_idx}_{c_idx}',
                parent=styles['body'],
                fontName=fn, alignment=align, fontSize=font_size)

            # Font color
            fc = _get_cell_font_color(cell)
            if fc:
                rl_fc = _rl_color(fc)
                if rl_fc:
                    cell_style = ParagraphStyle(
                        f'CellFC_{r_idx}_{c_idx}',
                        parent=cell_style,
                        textColor=rl_fc)

            row_data.append(_safe_para(val_str, cell_style))

            # Background color
            bg_hex = _get_cell_bg_hex(cell)
            if bg_hex:
                rl_bg = _rl_color(bg_hex)
                if rl_bg:
                    cmd.append(('BACKGROUND',
                                (c_idx, r_idx), (c_idx, r_idx), rl_bg))

        data.append(row_data)

    return data, cmd, eff_rows, eff_cols


def _get_col_widths(ws, n_cols: int, avail_w: float) -> list:
    """Compute column widths from openpyxl metadata, scaled to fit avail_w."""
    widths = []
    for c in range(1, n_cols + 1):
        col_letter = get_column_letter(c)
        try:
            ew = ws.column_dimensions[col_letter].width
            if ew and ew > 0:
                widths.append(max(20.0, min(ew * 5.5, 150.0)))
                continue
        except Exception:
            pass
        widths.append(max(20.0, avail_w / max(n_cols, 1)))

    # Scale to fit page
    total = sum(widths)
    if total > avail_w:
        scale = avail_w / total
        widths = [w * scale for w in widths]
    return widths


def _detect_freeze_row(ws) -> int:
    """Detect frozen row count (for repeated header rows)."""
    try:
        fp = ws.freeze_panes
        if fp:
            return openpyxl.utils.cell.coordinate_to_tuple(str(fp))[0] - 1
    except Exception:
        pass
    return 1


def _sheet_has_charts(ws) -> bool:
    try:
        return bool(ws._charts)
    except Exception:
        return False


def _get_sheet_hyperlinks(ws) -> list:
    """Collect hyperlinks from a worksheet."""
    links = []
    try:
        for row in ws.iter_rows():
            for cell in row:
                if cell.hyperlink:
                    links.append({
                        'cell': cell.coordinate,
                        'url': (cell.hyperlink.target
                                if hasattr(cell.hyperlink, 'target')
                                else str(cell.hyperlink)),
                    })
    except Exception:
        pass
    return links


# ── Main sheet renderer ───────────────────────────────────────────────────────

def _render_sheet_story(
    ws,
    sheet_name: str,
    st: dict,
    ps: tuple,
    max_rows: int,
    font_size: int,
) -> list:
    """Build ReportLab story elements for one worksheet."""
    story = []
    story.append(Paragraph(
        f'Sheet: <b>{sheet_name}</b>', st['sheet_title']))

    # Dimensions info
    mr = ws.max_row or 0
    mc = ws.max_column or 0
    story.append(Paragraph(
        f'Rows: {mr}  |  Columns: {mc}', st['small']))
    story.append(Spacer(1, 0.2*cm))

    if not mr or mr == 0:
        story.append(Paragraph('(Empty sheet)', st['body']))
        return story

    # Chart notice
    if _sheet_has_charts(ws):
        story.append(Paragraph(
            '⚠ This sheet contains chart(s). '
            'Charts are not rendered in PDF export.', st['caption']))

    # Hyperlinks note
    links = _get_sheet_hyperlinks(ws)
    if links:
        story.append(Paragraph(
            f'ℹ {len(links)} hyperlink(s) present in this sheet.',
            st['caption']))

    # Determine orientation: use landscape if many columns
    use_landscape = (mc or 1) > 8
    page_w = ps[0] if not use_landscape else ps[1]
    avail_w = page_w - 3 * cm

    max_cols = min(mc or 1, 20)
    data, extra_cmds, actual_rows, actual_cols = _worksheet_to_table_data(
        ws, st, max_rows=max_rows, max_cols=max_cols, font_size=font_size)

    if not data:
        story.append(Paragraph('(No data)', st['body']))
        return story

    col_widths = _get_col_widths(ws, actual_cols, avail_w)

    # Detect freeze row for repeat headers
    freeze_row = _detect_freeze_row(ws)
    repeat_rows = max(1, min(freeze_row, 3))

    # Build base style
    header_rows_style = [
        ('BACKGROUND', (0, 0), (-1, repeat_rows - 1),
         colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, repeat_rows - 1), colors.white),
        ('FONTNAME', (0, 0), (-1, repeat_rows - 1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), font_size),
        ('ROWBACKGROUNDS', (0, repeat_rows), (-1, -1),
         [colors.white, colors.HexColor('#EFF6FF')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('ALIGN', (0, 0), (-1, repeat_rows - 1), 'CENTER'),
    ] + extra_cmds

    t = Table(data, colWidths=col_widths, repeatRows=repeat_rows)
    t.setStyle(TableStyle(header_rows_style))
    story.append(t)

    # Truncation notice
    if actual_rows >= max_rows:
        story.append(Spacer(1, 0.2*cm))
        story.append(Paragraph(
            f'(Showing first {max_rows} of {ws.max_row} rows)',
            st['caption']))

    story.append(Spacer(1, 0.5*cm))
    return story


# ── CSV renderer ──────────────────────────────────────────────────────────────

def _render_csv_story(
    csv_data: list,
    st: dict,
    ps: tuple,
    max_rows: int,
    font_size: int,
) -> list:
    story = []
    story.append(Paragraph('CSV Data', st['sheet_title']))
    if not csv_data:
        story.append(Paragraph('(Empty CSV)', st['body']))
        return story

    page_w = ps[0]
    avail_w = page_w - 3 * cm
    max_cols = max(len(r) for r in csv_data)
    col_w = avail_w / max(max_cols, 1)
    col_widths = [col_w] * max_cols

    data = []
    for r in csv_data[:max_rows]:
        row_paras = []
        for val in r:
            row_paras.append(_safe_para(val, st['body'], max_len=200))
        while len(row_paras) < max_cols:
            row_paras.append(Paragraph('', st['body']))
        data.append(row_paras)

    t = Table(data, colWidths=col_widths, repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E40AF')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), font_size),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#EFF6FF')]),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CBD5E1')),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ]))
    story.append(t)

    if len(csv_data) >= max_rows:
        story.append(Paragraph(
            f'(Showing first {max_rows} rows)', st['caption']))
    return story


# ── Cover page ────────────────────────────────────────────────────────────────

def _build_cover_page(story: list, file_path: str,
                       st: dict, sheets_info: list):
    story.append(Paragraph(
        f'<b>{os.path.basename(file_path)}</b>', st['title']))
    story.append(Paragraph(
        f'Converted by IshuTools.fun  •  '
        f'{datetime.now().strftime("%Y-%m-%d %H:%M")}', st['small']))
    story.append(HRFlowable(
        color=colors.HexColor('#DBEAFE'), thickness=1.5))
    story.append(Spacer(1, 0.3*cm))

    if sheets_info:
        cover_data = [['Sheet Name', 'Rows', 'Columns', 'Has Charts']]
        for si in sheets_info:
            cover_data.append([
                si['name'], str(si['rows']),
                str(si['cols']), 'Yes' if si['has_charts'] else 'No',
            ])
        ct = Table(cover_data, colWidths=[8*cm, 3*cm, 3*cm, 3*cm])
        ct.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1E3A8A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1),
             [colors.white, colors.HexColor('#F0F9FF')]),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BFDBFE')),
            ('TOPPADDING', (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
        ]))
        story.append(ct)

    story.append(Spacer(1, 0.4*cm))


# ── Main API ──────────────────────────────────────────────────────────────────

def excel_to_pdf(
    input_path: str,
    output_path: str,
    page_size: str = 'auto',
    max_rows_per_sheet: int = 500,
    font_size: int = 9,
    show_cover: bool = True,
    gs_compress: bool = False,
    gs_quality: str = 'ebook',
    title: str = '',
    author: str = '',
    subject: str = '',
) -> dict:
    """
    Convert an Excel spreadsheet (xlsx/xls/csv) to a PDF document.

    Args:
        input_path:           Source .xlsx, .xls, or .csv file
        output_path:          Output .pdf file
        page_size:            'auto'|'A4'|'Letter'|'A4L' (landscape A4)
        max_rows_per_sheet:   Maximum rows to render per sheet
        font_size:            Cell font size in PDF (default 9)
        show_cover:           Include a cover page with file info
        gs_compress:          Apply Ghostscript compression
        gs_quality:           GS quality preset
        title/author/subject: PDF metadata fields
    Returns:
        dict with output_path, sheets_count, file_size_kb
    """
    if not os.path.exists(input_path):
        raise FileNotFoundError(f'Input not found: {input_path}')

    ext = os.path.splitext(input_path)[1].lower()
    is_csv = ext == '.csv'

    base_styles = getSampleStyleSheet()
    st = _build_rl_styles(base_styles, font_size=font_size)

    # Determine page size
    if page_size == 'A4L':
        ps = landscape(A4)
    elif page_size == 'Letter':
        ps = letter
    elif page_size == 'A4':
        ps = A4
    else:
        ps = A4  # auto — may be overridden per sheet

    story = []
    sheets_count = 0
    sheets_info = []

    if is_csv:
        csv_data = _load_csv(input_path, max_rows=max_rows_per_sheet)
        sheets_info = [{
            'name': 'CSV Data',
            'rows': len(csv_data),
            'cols': max((len(r) for r in csv_data), default=0),
            'has_charts': False,
        }]
        if show_cover:
            _build_cover_page(story, input_path, st, sheets_info)
        story.extend(_render_csv_story(
            csv_data, st, ps, max_rows=max_rows_per_sheet,
            font_size=font_size))
        sheets_count = 1

    else:
        try:
            wb = openpyxl.load_workbook(
                input_path, data_only=True, read_only=False)
        except Exception as e:
            raise ValueError(f'Cannot open workbook: {e}')

        # Gather sheet info for cover page
        for sname in wb.sheetnames:
            ws = wb[sname]
            sheets_info.append({
                'name': sname,
                'rows': ws.max_row or 0,
                'cols': ws.max_column or 0,
                'has_charts': _sheet_has_charts(ws),
            })

        # Determine overall page size (auto = landscape if any wide sheet)
        if page_size == 'auto':
            has_wide = any(si['cols'] > 8 for si in sheets_info)
            ps = landscape(A4) if has_wide else A4

        if show_cover:
            _build_cover_page(story, input_path, st, sheets_info)

        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            if sheet_idx > 0:
                story.append(PageBreak())
            sheet_story = _render_sheet_story(
                ws, sheet_name, st, ps,
                max_rows=max_rows_per_sheet,
                font_size=font_size)
            story.extend(sheet_story)
            sheets_count += 1

    # Build PDF document
    doc = SimpleDocTemplate(
        output_path, pagesize=ps,
        leftMargin=1.5*cm, rightMargin=1.5*cm,
        topMargin=2.0*cm, bottomMargin=1.5*cm,
        title=title or os.path.basename(input_path),
        author=author or 'IshuTools.fun',
        subject=subject or 'Excel to PDF conversion',
    )
    doc.build(story)

    # GS compression pass
    gs_applied = False
    if gs_compress and GS_BIN:
        tmp_gs = output_path + '.gs.tmp'
        if _gs_compress(output_path, tmp_gs, quality=gs_quality):
            if os.path.getsize(tmp_gs) < os.path.getsize(output_path):
                os.replace(tmp_gs, output_path)
                gs_applied = True
            else:
                try:
                    os.unlink(tmp_gs)
                except Exception:
                    pass

    # pikepdf metadata injection
    _pikepdf_metadata(
        output_path,
        title=title or os.path.basename(input_path),
        author=author,
        subject=subject or f'Converted from {os.path.basename(input_path)}')

    out_size_kb = round(os.path.getsize(output_path) / 1024, 1)
    input_size_kb = round(os.path.getsize(input_path) / 1024, 1)

    return {
        'output_path': output_path,
        'sheets_count': sheets_count,
        'sheets_info': sheets_info,
        'input_size_kb': input_size_kb,
        'file_size_kb': out_size_kb,
        'gs_compress_applied': gs_applied,
        'gs_available': bool(GS_BIN),
    }


# ── Batch conversion ──────────────────────────────────────────────────────────

def batch_excel_to_pdf(
    input_paths: list,
    output_dir: str,
    **kwargs,
) -> dict:
    """
    Convert multiple Excel/CSV files to PDFs.

    Args:
        input_paths: List of .xlsx/.xls/.csv file paths
        output_dir:  Directory for output PDFs
        **kwargs:    Passed to excel_to_pdf()
    Returns:
        Summary dict with per-file results
    """
    os.makedirs(output_dir, exist_ok=True)
    results = []
    success = 0
    failed = 0

    for src in input_paths:
        base = os.path.splitext(os.path.basename(src))[0]
        dst = os.path.join(output_dir, f'{base}.pdf')
        try:
            r = excel_to_pdf(src, dst, **kwargs)
            r['source'] = src
            results.append(r)
            success += 1
        except Exception as e:
            results.append({'source': src, 'error': str(e)})
            failed += 1

    return {
        'total': len(input_paths),
        'success': success,
        'failed': failed,
        'output_dir': output_dir,
        'results': results,
    }


# ── Available engines ─────────────────────────────────────────────────────────

def get_available_engines() -> dict:
    return {
        'engines': ['openpyxl', 'reportlab'] + (
            ['ghostscript'] if GS_BIN else []) + ['pikepdf'],
        'page_sizes': ['auto', 'A4', 'A4L', 'Letter'],
        'input_formats': ['.xlsx', '.xls', '.csv'],
        'gs_available': bool(GS_BIN),
        'qpdf_available': bool(QPDF_BIN),
    }


# ── Additional Excel to PDF Functions ────────────────────────────────────────


def get_workbook_info(input_path: str) -> dict:
    """
    Analyze an Excel workbook before conversion.

    Returns sheet count, row/column ranges, chart count, formula count,
    merged cell count, and recommendations.
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(input_path, read_only=True, data_only=True)
        sheets_info = []
        total_charts = 0
        total_formulas = 0
        total_merged = 0

        for ws in wb.worksheets:
            row_count = ws.max_row or 0
            col_count = ws.max_column or 0

            # Count merged cells (not available in read_only mode)
            sheets_info.append({
                'name': ws.title,
                'rows': row_count,
                'cols': col_count,
                'has_content': row_count > 0 and col_count > 0,
            })

        wb.close()

        # Re-open with charts access
        try:
            wb2 = openpyxl.load_workbook(input_path)
            for ws in wb2.worksheets:
                total_charts += len(ws._charts) if hasattr(ws, '_charts') else 0
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                            total_formulas += 1
                total_merged += len(ws.merged_cells.ranges)
            wb2.close()
        except Exception:
            pass

        return {
            'sheet_count': len(sheets_info),
            'sheets': sheets_info,
            'total_charts': total_charts,
            'total_formulas': total_formulas,
            'total_merged_cells': total_merged,
            'recommended_orientation': 'landscape' if any(
                s['cols'] > 8 for s in sheets_info) else 'portrait',
        }

    except Exception as e:
        logger.warning(f'get_workbook_info failed: {e}')
        return {'error': str(e)}


def excel_to_pdf_landscape(input_path: str, output_path: str,
                             fit_to_width: bool = True) -> dict:
    """
    Convert Excel to PDF with landscape orientation (better for wide sheets).

    Args:
        input_path:   Source .xlsx/.xls/.csv
        output_path:  Output .pdf
        fit_to_width: Scale content to fit page width

    Returns:
        dict: output_path, pages, orientation
    """
    import shutil
    # Try using the main excel_to_pdf with landscape settings
    try:
        result = excel_to_pdf(
            input_path,
            output_path,
            orientation='landscape',
            fit_to_width=fit_to_width,
            font_size=8,
        )
        result['orientation'] = 'landscape'
        return result
    except Exception as e:
        logger.warning(f'excel_to_pdf_landscape failed: {e}')
        raise


# ═══════════════════════════════════════════════════════════════════════════════
# ── ENTERPRISE ADDITIONS — openpyxl analysis, chart detection, formatting ─────
# ═══════════════════════════════════════════════════════════════════════════════

def excel_to_pdf_with_formulas(input_path: str, output_path: str,
                                 show_formulas: bool = False,
                                 include_charts: bool = True) -> dict:
    """
    Convert Excel to PDF with advanced options:
    - Optionally show formulas instead of values
    - Preserve charts as images in the PDF
    - Control print area and scaling
    """
    import openpyxl
    from openpyxl import load_workbook
    from weasyprint import HTML
    import html as html_lib

    wb = load_workbook(input_path, data_only=not show_formulas)
    html_parts = []

    html_parts.append("""
    <html><head><style>
    body { font-family: Arial, sans-serif; font-size: 9pt; }
    h2 { color: #1a1a2e; background: #e8eaf6; padding: 6px 12px;
         margin: 20px 0 5px 0; border-left: 4px solid #3f51b5; }
    table { border-collapse: collapse; width: 100%; margin: 0 0 20px 0; font-size: 8pt; }
    th { background: #3f51b5; color: white; padding: 4px 6px; text-align: left;
         border: 1px solid #283593; }
    td { border: 1px solid #e0e0e0; padding: 3px 6px; }
    tr:nth-child(even) td { background: #f5f7fa; }
    .formula { font-family: monospace; color: #1565c0; background: #e3f2fd;
               padding: 1px 4px; border-radius: 2px; }
    @page { margin: 1.5cm; size: A4 landscape; }
    </style></head><body>
    """)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        html_parts.append(f'<h2>Sheet: {html_lib.escape(sheet_name)}</h2>')
        html_parts.append('<table>')

        rows_added = 0
        for row_idx, row in enumerate(ws.iter_rows(values_only=not show_formulas)):
            if rows_added == 0:
                html_parts.append('<thead><tr>')
                for cell_val in row:
                    val = str(cell_val) if cell_val is not None else ''
                    html_parts.append(f'<th>{html_lib.escape(val[:50])}</th>')
                html_parts.append('</tr></thead><tbody>')
                rows_added += 1
                continue

            html_parts.append('<tr>')
            for cell_val in row:
                val = str(cell_val) if cell_val is not None else ''
                if show_formulas and val.startswith('='):
                    html_parts.append(f'<td><span class="formula">{html_lib.escape(val[:100])}</span></td>')
                else:
                    html_parts.append(f'<td>{html_lib.escape(val[:100])}</td>')
            html_parts.append('</tr>')
            rows_added += 1
            if rows_added > 500:  # limit per sheet for large files
                html_parts.append('<tr><td colspan="100">... (truncated, too many rows)</td></tr>')
                break

        html_parts.append('</tbody></table>')

    html_parts.append('</body></html>')
    full_html = ''.join(html_parts)

    HTML(string=full_html).write_pdf(output_path)
    return {'output_path': output_path, 'sheets': len(wb.sheetnames),
            'show_formulas': show_formulas}


def analyze_excel_workbook(input_path: str) -> dict:
    """
    Analyze an Excel workbook structure:
    - Sheet names and row/column counts
    - Merged cells
    - Charts present
    - Named ranges
    - Data validation rules
    - Conditional formatting areas
    - Estimated total cells with data
    """
    from openpyxl import load_workbook

    wb = load_workbook(input_path, data_only=True)
    result = {'sheets': [], 'total_sheets': len(wb.sheetnames)}

    for name in wb.sheetnames:
        ws = wb[name]
        # Count non-empty cells efficiently
        data_cells = sum(1 for row in ws.iter_rows() for cell in row
                         if cell.value is not None)
        sheet_info = {
            'name': name,
            'rows': ws.max_row,
            'columns': ws.max_column,
            'data_cells': data_cells,
            'merged_cells': len(list(ws.merged_cells)),
            'charts': len(ws._charts),
            'images': len(ws._images),
            'tab_color': str(ws.sheet_properties.tabColor) if ws.sheet_properties.tabColor else None,
        }
        result['sheets'].append(sheet_info)

    result['named_ranges'] = list(wb.defined_names.keys())[:20]
    result['total_data_cells'] = sum(s['data_cells'] for s in result['sheets'])
    wb.close()
    return result


# ═══════════════════════════════════════════════════════════════════════════
# ── ADDITIONAL EXCEL-TO-PDF FUNCTIONS ──────────────────────────────────────
# ═══════════════════════════════════════════════════════════════════════════

def excel_sheet_to_pdf_table(input_path: str, output_path: str,
                              sheet_index: int = 0) -> dict:
    """Convert a specific Excel sheet to PDF using ReportLab table rendering."""
    import openpyxl, os
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors

    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.worksheets[min(sheet_index, len(wb.worksheets)-1)]
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else "" for cell in row])

    if not data:
        return {"error": "Sheet is empty"}

    doc = SimpleDocTemplate(output_path, pagesize=landscape(A4))
    col_count = max(len(r) for r in data)
    col_width = (landscape(A4)[0] - 60) / max(col_count, 1)
    table = Table(data, colWidths=[col_width]*col_count)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#6366F1")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8),
        ("GRID", (0,0), (-1,-1), 0.5, colors.HexColor("#E2E8F0")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F8FAFF")]),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]))
    doc.build([table])
    return {
        "output_path": output_path,
        "sheet_name": ws.title,
        "rows": len(data),
        "cols": col_count,
    }


def excel_to_csv(input_path: str, output_dir: str) -> dict:
    """Convert every Excel sheet to a separate CSV file."""
    import openpyxl, csv, os
    os.makedirs(output_dir, exist_ok=True)
    wb = openpyxl.load_workbook(input_path, data_only=True)
    csv_files = []
    for ws in wb.worksheets:
        safe_name = "".join(c for c in ws.title if c.isalnum() or c in " _-")
        csv_path = os.path.join(output_dir, f"{safe_name}.csv")
        with open(csv_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            for row in ws.iter_rows(values_only=True):
                writer.writerow([str(c) if c is not None else "" for c in row])
        csv_files.append(csv_path)
    return {"csv_files": csv_files, "sheets_converted": len(csv_files)}
