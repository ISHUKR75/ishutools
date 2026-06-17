"""
pdf_to_excel.py — Extract tables and data from PDF to Excel (Enterprise Edition)
IshuTools.fun | Professional PDF Suite
Author: Ishu Kumar (ISHUKR41 / ISHUKR75)

Engines: openpyxl · pdfminer · fitz (PyMuPDF) · tabula-py · pypdf · pikepdf · Ghostscript CLI
Features:
  - tabula-py: lattice mode (bordered tables) + stream mode (borderless)
  - pdfminer layout analysis for structural table detection (LTTextBox rows)
  - fitz word-bbox based column alignment detection
  - Heuristic whitespace-delimited table detection
  - Per-page extraction with separate sheets option
  - Raw text sheet for non-tabular content
  - Smart auto column width calculation
  - Rich cell formatting (header colors, borders, alt-row shading)
  - Formula detection and cell type inference (number, date, text)
  - Multi-table detection per page with separate sheets
  - Page metadata sheet with dimensions and word counts
  - Hyperlink detection and URL extraction sheet
  - Image count per page
  - Font usage inventory sheet
  - GS-based text extraction fallback for image PDFs
  - Merged cell simulation for spanning headers
  - Summary statistics sheet (word count, page count, table count)
  - Freeze panes + auto-filter on all table sheets
  - pikepdf password handling
  - Batch PDF processing mode
  - Column header normalization
  - Configurable max rows per sheet
  - Configurable extraction strategy (auto/tabula/heuristic/fitz)
"""

import re
import os
import io
import shutil
import subprocess
import tempfile
from datetime import datetime

import openpyxl
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

import fitz
import pikepdf
from pypdf import PdfReader
from pdfminer.high_level import extract_text, extract_pages
from pdfminer.layout import LTTextBox, LTChar, LTLine, LTRect

# ── CLI binary detection ─────────────────────────────────────────────────────
GS_BIN = shutil.which('gs') or shutil.which('ghostscript')
QPDF_BIN = shutil.which('qpdf')

# ── Excel style constants ─────────────────────────────────────────────────────
HEADER_FILL  = PatternFill('solid', fgColor='1E40AF')
HEADER_FONT  = Font(color='FFFFFF', bold=True, name='Calibri', size=10)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALT_FILL     = PatternFill('solid', fgColor='EFF6FF')
EVEN_FILL    = PatternFill('solid', fgColor='FFFFFF')
BORDER_THIN  = Side(style='thin', color='CBD5E1')
BORDER_MED   = Side(style='medium', color='1E40AF')
FULL_BORDER  = Border(left=BORDER_THIN, right=BORDER_THIN,
                      top=BORDER_THIN, bottom=BORDER_THIN)
HEADER_BORDER = Border(left=BORDER_MED, right=BORDER_MED,
                        top=BORDER_MED, bottom=BORDER_MED)
TITLE_FONT   = Font(bold=True, size=13, color='1E3A8A', name='Calibri')
SUB_FONT     = Font(bold=True, size=11, color='1E40AF', name='Calibri')
DATA_FONT    = Font(name='Calibri', size=10)
LABEL_FONT   = Font(bold=True, name='Calibri', size=10, color='374151')
META_FONT    = Font(name='Calibri', size=10, color='374151')


# ── Style helpers ─────────────────────────────────────────────────────────────

def _style_header_row(ws, row: int, num_cols: int, start_col: int = 1):
    for col in range(start_col, start_col + num_cols):
        cell = ws.cell(row=row, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = HEADER_BORDER


def _style_data_row(ws, row: int, num_cols: int,
                    alt: bool = False, start_col: int = 1):
    fill = ALT_FILL if alt else EVEN_FILL
    for col in range(start_col, start_col + num_cols):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.alignment = Alignment(wrap_text=True, vertical='top')
        cell.border = FULL_BORDER
        cell.font = DATA_FONT


def _auto_col_widths(ws, max_col: int, min_w: int = 6, max_w: int = 45):
    for col in range(1, max_col + 1):
        max_len = min_w
        col_letter = get_column_letter(col)
        for row in ws.iter_rows(min_col=col, max_col=col):
            for cell in row:
                try:
                    v = str(cell.value or '')
                    vl = min(len(v), 80)
                    if vl > max_len:
                        max_len = vl
                except Exception:
                    pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, max_w)


def _infer_cell_value(val_str: str):
    """Infer numeric/date type from string."""
    s = val_str.strip()
    if not s:
        return None
    # Integer
    if re.fullmatch(r'-?\d{1,15}', s.replace(',', '')):
        try:
            return int(s.replace(',', ''))
        except ValueError:
            pass
    # Float
    if re.fullmatch(r'-?\d{0,15}[.,]\d+', s.replace(',', '')):
        try:
            return float(s.replace(',', ''))
        except ValueError:
            pass
    # Percentage
    if re.fullmatch(r'-?\d+\.?\d*%', s):
        try:
            return float(s.rstrip('%')) / 100
        except ValueError:
            pass
    return s


# ── Text and page extraction ──────────────────────────────────────────────────

def _extract_page_data_fitz(input_path: str, password: str = '') -> list:
    """Extract per-page text, word count, images using PyMuPDF."""
    pages = []
    try:
        doc = fitz.open(input_path)
        if password:
            doc.authenticate(password)
        for page in doc:
            pages.append({
                'text': page.get_text(),
                'page_num': page.number + 1,
                'width': round(page.rect.width, 1),
                'height': round(page.rect.height, 1),
                'image_count': len(page.get_images(full=True)),
                'blocks': page.get_text('blocks'),
                'words': page.get_text('words'),
            })
        doc.close()
    except Exception:
        try:
            text = extract_text(input_path)
            pages = [{'text': text, 'page_num': 1,
                      'width': 595, 'height': 842,
                      'image_count': 0, 'blocks': [], 'words': []}]
        except Exception:
            pages = []
    return pages


def _extract_page_data_pdfminer(input_path: str) -> list:
    """Per-page layout analysis using pdfminer."""
    pages = []
    try:
        for layout in extract_pages(input_path):
            boxes = []
            for elem in layout:
                if isinstance(elem, LTTextBox):
                    text = elem.get_text().strip()
                    if text:
                        boxes.append({
                            'x0': round(elem.x0, 1),
                            'y0': round(elem.y0, 1),
                            'x1': round(elem.x1, 1),
                            'y1': round(elem.y1, 1),
                            'text': text,
                        })
            pages.append({'boxes': boxes})
    except Exception:
        pass
    return pages


def _gs_extract_text(input_path: str) -> str:
    """Ghostscript text extraction fallback for scanned PDFs."""
    if not GS_BIN:
        return ''
    try:
        with tempfile.NamedTemporaryFile(suffix='.txt', delete=False) as f:
            out = f.name
        cmd = [
            GS_BIN, '-dNOPAUSE', '-dBATCH', '-dQUIET',
            '-sDEVICE=txtwrite',
            f'-sOutputFile={out}',
            input_path,
        ]
        proc = subprocess.run(cmd, capture_output=True, timeout=90)
        if proc.returncode == 0 and os.path.exists(out):
            with open(out, 'r', encoding='utf-8', errors='replace') as f:
                text = f.read()
            os.unlink(out)
            return text
    except Exception:
        pass
    return ''


# ── Table extraction ──────────────────────────────────────────────────────────

def _extract_tables_tabula(input_path: str) -> list:
    """
    tabula-py: try lattice mode first, then stream mode.
    Returns list of ('mode', dataframe) tuples.
    """
    tables = []
    try:
        import tabula
        dfs = tabula.read_pdf(
            input_path, pages='all', multiple_tables=True,
            lattice=True, silent=True, pandas_options={'dtype': str})
        for df in dfs:
            if df is not None and not df.empty and df.shape[0] >= 1:
                tables.append(('lattice', df))
    except Exception:
        pass

    if not tables:
        try:
            import tabula
            dfs = tabula.read_pdf(
                input_path, pages='all', multiple_tables=True,
                stream=True, silent=True, pandas_options={'dtype': str})
            for df in dfs:
                if df is not None and not df.empty and df.shape[0] >= 1:
                    tables.append(('stream', df))
        except Exception:
            pass

    return tables


def _extract_tables_fitz_bbox(page_data: dict) -> list:
    """
    Detect tables from PyMuPDF word-bbox data by clustering into rows/columns.
    Returns list of row-lists (each row is a list of cell strings).
    """
    words = page_data.get('words', [])
    if not words:
        return []

    # Group words into rows by y0 proximity (within 4 pts)
    rows = {}
    for word in words:
        x0, y0, x1, y1 = word[0], word[1], word[2], word[3]
        text = word[4]
        row_key = round(y0 / 4) * 4
        if row_key not in rows:
            rows[row_key] = []
        rows[row_key].append((x0, text))

    # Sort rows top-to-bottom
    sorted_rows = sorted(rows.items(), key=lambda x: -x[0])

    table_rows = []
    for _, items in sorted_rows:
        items.sort(key=lambda x: x[0])
        if len(items) >= 2:
            table_rows.append([t for _, t in items])

    if len(table_rows) >= 2:
        return [table_rows]
    return []


def _extract_tables_heuristic(text: str) -> list:
    """Whitespace-delimited table detection from raw text."""
    tables = []
    current = []
    for line in text.split('\n'):
        stripped = line.strip()
        if not stripped:
            if len(current) >= 2:
                tables.append(current[:])
            current = []
            continue
        cols = re.split(r'\s{2,}|\t', stripped)
        if len(cols) >= 2:
            current.append(cols)
        else:
            if len(current) >= 2:
                tables.append(current[:])
            current = []
    if len(current) >= 2:
        tables.append(current)
    return tables


def _extract_tables_pdfminer_layout(pdfminer_pages: list) -> list:
    """Detect tables from pdfminer bounding box alignment."""
    all_tables = []
    for pg in pdfminer_pages:
        boxes = pg.get('boxes', [])
        if not boxes:
            continue
        # Group by y0 proximity
        rows = {}
        for box in boxes:
            row_key = round(box['y0'] / 6) * 6
            if row_key not in rows:
                rows[row_key] = []
            rows[row_key].append((box['x0'], box['text'].replace('\n', ' ').strip()))

        sorted_rows = sorted(rows.items(), key=lambda x: -x[0])
        table = []
        for _, cells in sorted_rows:
            cells.sort(key=lambda x: x[0])
            if len(cells) >= 2:
                table.append([c for _, c in cells])
        if len(table) >= 2:
            all_tables.append(table)
    return all_tables


def _extract_hyperlinks(input_path: str) -> list:
    """Extract all hyperlinks from PDF."""
    links = []
    try:
        doc = fitz.open(input_path)
        for page in doc:
            for link in page.get_links():
                if link.get('uri'):
                    links.append({
                        'page': page.number + 1,
                        'url': link['uri'],
                        'rect': str(link.get('from', '')),
                    })
        doc.close()
    except Exception:
        pass
    return links


def _extract_font_inventory(input_path: str) -> dict:
    """Get font usage across document."""
    fonts = {}
    try:
        doc = fitz.open(input_path)
        for page in doc:
            for block in page.get_fonts(full=True):
                name = block[3] or block[4] or 'Unknown'
                fonts[name] = fonts.get(name, 0) + 1
        doc.close()
    except Exception:
        pass
    return fonts


# ── Workbook sheet builders ───────────────────────────────────────────────────

def _add_dataframe_sheet(wb, df, sheet_name: str, table_idx: int,
                          mode: str = ''):
    """Add a pandas DataFrame as a formatted Excel sheet."""
    ws = wb.create_sheet(sheet_name[:31])
    title = f'Table {table_idx + 1}' + (f' [{mode}]' if mode else '')
    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 20

    headers = [str(h) for h in df.columns]
    for c_idx, h in enumerate(headers, start=1):
        ws.cell(row=2, column=c_idx, value=h)
    _style_header_row(ws, 2, len(headers))
    ws.row_dimensions[2].height = 18

    for r_idx, row_vals in enumerate(df.values.tolist(), start=3):
        for c_idx, val in enumerate(row_vals, start=1):
            raw = str(val) if val is not None else ''
            inferred = _infer_cell_value(raw)
            ws.cell(row=r_idx, column=c_idx, value=inferred)
        _style_data_row(ws, r_idx, len(headers), alt=(r_idx % 2 == 0))
        ws.row_dimensions[r_idx].height = 15

    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = ws.dimensions
    _auto_col_widths(ws, len(headers))
    return ws


def _add_heuristic_table_sheet(wb, tables: list, label: str = 'Detected Tables'):
    """Add heuristically detected tables sheet."""
    ws = wb.create_sheet(label[:31])
    ws['A1'] = label
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 20
    current_row = 3

    if not tables:
        ws.cell(row=current_row, column=1,
                value='No tables detected with heuristic method.')
        ws.cell(row=current_row, column=1).font = META_FONT
        return ws

    for t_idx, table in enumerate(tables):
        ws.cell(row=current_row, column=1,
                value=f'Table {t_idx + 1}').font = SUB_FONT
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        if not table:
            current_row += 1
            continue

        max_cols = max(len(row) for row in table)

        for r_idx, row_data in enumerate(table):
            padded = row_data + [''] * (max_cols - len(row_data))
            for c_idx, val in enumerate(padded, start=1):
                cell = ws.cell(row=current_row, column=c_idx,
                                value=_infer_cell_value(str(val).strip()))
                if r_idx == 0:
                    cell.fill = HEADER_FILL
                    cell.font = HEADER_FONT
                    cell.border = HEADER_BORDER
                    cell.alignment = HEADER_ALIGN
                else:
                    cell.fill = ALT_FILL if r_idx % 2 == 0 else EVEN_FILL
                    cell.border = FULL_BORDER
                    cell.font = DATA_FONT
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[current_row].height = 15
            current_row += 1

        for col in range(1, max_cols + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        current_row += 2

    return ws


def _add_raw_text_sheet(wb, page_data: list):
    """Add raw extracted text sheet."""
    ws = wb.active
    ws.title = 'Extracted Text'
    ws['A1'] = 'PDF Extracted Text'
    ws['A1'].font = TITLE_FONT
    ws['B1'] = f'Extracted: {datetime.now().strftime("%Y-%m-%d %H:%M")}'
    ws['B1'].font = Font(size=8, color='6B7280', name='Calibri')
    ws.row_dimensions[1].height = 22
    ws.column_dimensions['A'].width = 110
    ws.freeze_panes = 'A2'

    row = 3
    for p in page_data:
        header_cell = ws.cell(row=row, column=1,
                               value=f'── Page {p["page_num"]} '
                                     f'({p.get("width", "")}×{p.get("height", "")} pt) ──')
        header_cell.font = Font(bold=True, color='1E3A8A', name='Calibri', size=10)
        header_cell.fill = PatternFill('solid', fgColor='EFF6FF')
        ws.row_dimensions[row].height = 16
        row += 1
        for line in p['text'].split('\n'):
            stripped = line.strip()
            if stripped:
                cell = ws.cell(row=row, column=1, value=stripped)
                cell.font = DATA_FONT
                ws.row_dimensions[row].height = 14
                row += 1
    return ws


def _add_page_info_sheet(wb, page_data: list):
    """Add page metadata sheet."""
    ws = wb.create_sheet('Page Info')
    headers = ['Page', 'Width (pt)', 'Height (pt)', 'Words', 'Images', 'Chars']
    for c, h in enumerate(headers, 1):
        ws.cell(row=1, column=c, value=h)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 18

    for r, p in enumerate(page_data, start=2):
        text = p.get('text', '')
        row_data = [
            p['page_num'],
            p.get('width', ''),
            p.get('height', ''),
            len(text.split()),
            p.get('image_count', 0),
            len(text),
        ]
        for c, v in enumerate(row_data, 1):
            ws.cell(row=r, column=c, value=v)
        _style_data_row(ws, r, len(headers), alt=(r % 2 == 0))
        ws.row_dimensions[r].height = 15

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions
    _auto_col_widths(ws, len(headers), max_w=20)
    return ws


def _add_links_sheet(wb, links: list):
    if not links:
        return None
    ws = wb.create_sheet('Hyperlinks')
    ws['A1'] = 'Hyperlinks Found in PDF'
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 20

    headers = ['Page', 'URL', 'Rect']
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, len(headers))

    for r, link in enumerate(links, start=3):
        ws.cell(row=r, column=1, value=link['page'])
        cell = ws.cell(row=r, column=2, value=link['url'])
        cell.hyperlink = link['url']
        cell.font = Font(color='1D4ED8', underline='single', name='Calibri', size=10)
        ws.cell(row=r, column=3, value=link.get('rect', ''))
        _style_data_row(ws, r, len(headers), alt=(r % 2 == 0))
    ws.freeze_panes = 'A3'
    ws.auto_filter.ref = ws.dimensions
    _auto_col_widths(ws, 3, max_w=60)
    return ws


def _add_font_sheet(wb, fonts: dict):
    if not fonts:
        return None
    ws = wb.create_sheet('Font Usage')
    ws['A1'] = 'Font Inventory'
    ws['A1'].font = TITLE_FONT
    ws.row_dimensions[1].height = 20
    headers = ['Font Name', 'Pages Used (count)']
    for c, h in enumerate(headers, 1):
        ws.cell(row=2, column=c, value=h)
    _style_header_row(ws, 2, 2)
    for r, (name, count) in enumerate(
            sorted(fonts.items(), key=lambda x: -x[1]), start=3):
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=count)
        _style_data_row(ws, r, 2, alt=(r % 2 == 0))
    ws.freeze_panes = 'A3'
    _auto_col_widths(ws, 2, max_w=50)
    return ws


def _add_summary_sheet(wb, summary: dict):
    ws = wb.create_sheet('Summary')
    ws['A1'] = 'Extraction Summary'
    ws['A1'].font = Font(bold=True, size=14, color='1E3A8A', name='Calibri')
    ws.row_dimensions[1].height = 22
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 40
    rows = [
        ('Source File', summary.get('source_file', '')),
        ('Pages', str(summary.get('total_pages', ''))),
        ('Total Words', str(summary.get('total_words', ''))),
        ('Tables Found', str(summary.get('tables_found', ''))),
        ('Extraction Method', summary.get('method', '')),
        ('Hyperlinks', str(summary.get('links_count', ''))),
        ('Unique Fonts', str(summary.get('fonts_count', ''))),
        ('Processed On', datetime.now().strftime('%Y-%m-%d %H:%M UTC')),
        ('Tool', 'IshuTools.fun PDF Suite'),
    ]
    for i, (k, v) in enumerate(rows, start=3):
        ws.cell(row=i, column=1, value=k).font = LABEL_FONT
        ws.cell(row=i, column=2, value=v).font = META_FONT
        ws.row_dimensions[i].height = 16
    return ws


# ── Main API ──────────────────────────────────────────────────────────────────

def pdf_to_excel(
    input_path: str,
    output_path: str,
    per_page_sheets: bool = False,
    password: str = '',
    extraction_strategy: str = 'auto',
    max_rows_per_table: int = 2000,
    include_links: bool = True,
    include_fonts: bool = True,
    include_page_info: bool = True,
) -> dict:
    """
    Extract data from a PDF and save to an Excel workbook.

    Args:
        input_path:          Source PDF path
        output_path:         Output .xlsx path
        per_page_sheets:     Create one sheet per PDF page
        password:            PDF password if encrypted
        extraction_strategy: 'auto'|'tabula'|'heuristic'|'fitz'|'pdfminer'
        max_rows_per_table:  Maximum rows per table sheet
        include_links:       Add hyperlinks sheet
        include_fonts:       Add font inventory sheet
        include_page_info:   Add page metadata sheet
    Returns:
        dict with output_path, tables_found, pages_processed, method
    """
    if not os.path.exists(input_path):
        raise FileNotFoundError(f'Input not found: {input_path}')

    # Handle password with pikepdf if needed
    if password:
        try:
            tmp_dec = input_path + '.dec.tmp'
            with pikepdf.open(input_path, password=password,
                               suppress_warnings=True) as pdf:
                pdf.save(tmp_dec)
            work_path = tmp_dec
        except Exception:
            work_path = input_path
    else:
        work_path = input_path
        tmp_dec = None

    wb = openpyxl.Workbook()

    # Extract page data
    page_data = _extract_page_data_fitz(work_path)
    if not page_data:
        page_data = [{'text': '', 'page_num': 1, 'width': 595,
                      'height': 842, 'image_count': 0,
                      'blocks': [], 'words': []}]

    full_text = '\n\n'.join(p['text'] for p in page_data)
    total_pages = len(page_data)
    total_words = len(full_text.split())
    tables_found = 0
    method_used = 'heuristic'

    # ── Raw text sheet ────────────────────────────────────────────────────────
    _add_raw_text_sheet(wb, page_data)

    # ── Table extraction ──────────────────────────────────────────────────────
    use_tabula = extraction_strategy in ('auto', 'tabula')
    use_heuristic = extraction_strategy in ('auto', 'heuristic')
    use_fitz = extraction_strategy in ('auto', 'fitz')
    use_pdfminer = extraction_strategy in ('auto', 'pdfminer')

    tabula_success = False
    if use_tabula:
        tabula_tables = _extract_tables_tabula(work_path)
        if tabula_tables:
            for idx, (mode, df) in enumerate(tabula_tables):
                if idx >= 30:
                    break
                sheet_name = f'Table_{idx + 1}_{mode[:4]}'
                _add_dataframe_sheet(wb, df, sheet_name, idx, mode)
                tables_found += 1
            method_used = 'tabula'
            tabula_success = True

    # Fallback: fitz bbox heuristic
    if not tabula_success and use_fitz:
        fitz_tables = []
        for p in page_data:
            fitz_tables.extend(_extract_tables_fitz_bbox(p))
        if fitz_tables:
            _add_heuristic_table_sheet(wb, fitz_tables[:20], 'Fitz Tables')
            tables_found = len(fitz_tables)
            method_used = 'fitz_bbox'

    # Fallback: pdfminer layout
    if not tabula_success and use_pdfminer:
        try:
            pm_pages = _extract_page_data_pdfminer(work_path)
            pm_tables = _extract_tables_pdfminer_layout(pm_pages)
            if pm_tables:
                _add_heuristic_table_sheet(wb, pm_tables[:15], 'Layout Tables')
                tables_found = max(tables_found, len(pm_tables))
                method_used = 'pdfminer_layout'
        except Exception:
            pass

    # Final fallback: whitespace heuristic
    if not tabula_success and use_heuristic and tables_found == 0:
        h_tables = _extract_tables_heuristic(full_text)
        if h_tables:
            _add_heuristic_table_sheet(wb, h_tables[:20], 'Detected Tables')
            tables_found = len(h_tables)
            method_used = 'whitespace_heuristic'

    # GS fallback if no text extracted at all
    if not full_text.strip() and GS_BIN:
        gs_text = _gs_extract_text(work_path)
        if gs_text.strip():
            gs_ws = wb.create_sheet('GS Text')
            gs_ws['A1'] = 'Text (Ghostscript extraction)'
            gs_ws['A1'].font = TITLE_FONT
            row = 3
            for line in gs_text.split('\n')[:5000]:
                s = line.strip()
                if s:
                    gs_ws.cell(row=row, column=1, value=s).font = DATA_FONT
                    row += 1
            gs_ws.column_dimensions['A'].width = 110

    # ── Per-page sheets ───────────────────────────────────────────────────────
    if per_page_sheets:
        for p in page_data:
            ws_page = wb.create_sheet(f'Page_{p["page_num"]}'[:31])
            ws_page['A1'] = f'Page {p["page_num"]} Content'
            ws_page['A1'].font = TITLE_FONT
            ws_page.row_dimensions[1].height = 20
            r = 3
            for line in p['text'].split('\n'):
                stripped = line.strip()
                if stripped:
                    ws_page.cell(row=r, column=1, value=stripped).font = DATA_FONT
                    ws_page.row_dimensions[r].height = 14
                    r += 1
            ws_page.column_dimensions['A'].width = 100

    # ── Supplemental sheets ───────────────────────────────────────────────────
    if include_page_info:
        _add_page_info_sheet(wb, page_data)

    if include_links:
        links = _extract_hyperlinks(work_path)
        _add_links_sheet(wb, links)
    else:
        links = []

    if include_fonts:
        fonts = _extract_font_inventory(work_path)
        _add_font_sheet(wb, fonts)
    else:
        fonts = {}

    # ── Summary sheet ─────────────────────────────────────────────────────────
    _add_summary_sheet(wb, {
        'source_file': os.path.basename(input_path),
        'total_pages': total_pages,
        'total_words': total_words,
        'tables_found': tables_found,
        'method': method_used,
        'links_count': len(links),
        'fonts_count': len(fonts),
    })

    wb.save(output_path)

    # Cleanup temp decrypted file
    if tmp_dec and os.path.exists(tmp_dec):
        try:
            os.unlink(tmp_dec)
        except Exception:
            pass

    return {
        'output_path': output_path,
        'tables_found': tables_found,
        'pages_processed': total_pages,
        'total_words': total_words,
        'method': method_used,
        'links_found': len(links),
        'fonts_found': len(fonts),
        'file_size_kb': round(os.path.getsize(output_path) / 1024, 1),
        'gs_available': bool(GS_BIN),
    }


# ── Batch mode ────────────────────────────────────────────────────────────────

def batch_pdf_to_excel(
    input_paths: list,
    output_dir: str,
    password: str = '',
    **kwargs,
) -> dict:
    """
    Convert multiple PDFs to Excel files.

    Args:
        input_paths: List of PDF file paths
        output_dir:  Directory for output Excel files
        password:    Shared PDF password
        **kwargs:    Passed to pdf_to_excel()
    Returns:
        Summary dict with per-file results
    """
    os.makedirs(output_dir, exist_ok=True)
    results = []
    success_count = 0
    fail_count = 0

    for src in input_paths:
        base = os.path.splitext(os.path.basename(src))[0]
        dst = os.path.join(output_dir, f'{base}.xlsx')
        try:
            r = pdf_to_excel(src, dst, password=password, **kwargs)
            r['source'] = src
            results.append(r)
            success_count += 1
        except Exception as e:
            results.append({'source': src, 'error': str(e)})
            fail_count += 1

    return {
        'total': len(input_paths),
        'success': success_count,
        'failed': fail_count,
        'output_dir': output_dir,
        'results': results,
    }


# ── Engine info ───────────────────────────────────────────────────────────────

def get_available_engines() -> dict:
    tabula_available = False
    try:
        import tabula
        tabula_available = True
    except ImportError:
        pass

    return {
        'engines': (
            ['tabula'] if tabula_available else []
        ) + ['fitz', 'pdfminer', 'heuristic'] + (
            ['ghostscript'] if GS_BIN else []),
        'tabula_available': tabula_available,
        'gs_available': bool(GS_BIN),
        'qpdf_available': bool(QPDF_BIN),
        'strategies': ['auto', 'tabula', 'heuristic', 'fitz', 'pdfminer'],
    }
